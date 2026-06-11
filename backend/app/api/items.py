"""Items endpoints - query and manage tracked YouTube channels."""

import io
import json
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import case, delete, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.crawl_job import CrawlJob
from app.models.item import Item
from app.models.user import User
from app.schemas.item import ItemGroupSummary, ItemIdsResponse, ItemListResponse, ItemMoveRequest, ItemResponse, ItemSummaryResponse
from app.services.auth import get_current_user
from app.utils.youtube_urls import build_channel_url

router = APIRouter()


def _plain_query_value(value, default=None):
    return default if value.__class__.__name__ == "Query" else value


async def _visible_user_ids(db: AsyncSession, actor: User, requested_user_id: str | None = None) -> list[str]:
    if actor.role == "admin":
        if requested_user_id:
            return [requested_user_id]
        result = await db.execute(select(User.id))
        return [row[0] for row in result.all()]
    if actor.role == "manager":
        if requested_user_id:
            result = await db.execute(
                select(User).where(or_(User.id == actor.id, User.manager_id == actor.id), User.id == requested_user_id)
            )
            if not result.scalar_one_or_none():
                raise HTTPException(status_code=403, detail="Not authorized for this user")
            return [requested_user_id]
        result = await db.execute(select(User.id).where(or_(User.id == actor.id, User.manager_id == actor.id)))
        return [row[0] for row in result.all()]
    if requested_user_id and requested_user_id != actor.id:
        raise HTTPException(status_code=403, detail="Not authorized for this user")
    return [actor.id]


async def _user_map(db: AsyncSession, user_ids: list[str]) -> dict[str, User]:
    if not user_ids:
        return {}
    result = await db.execute(select(User).where(User.id.in_(user_ids)))
    return {user.id: user for user in result.scalars().all()}


def _format_added_date(dt: datetime | None) -> str | None:
    return dt.strftime("%d/%m %H:%M") if dt else None


def _response(item: Item, user: User | None = None) -> ItemResponse:
    url = build_channel_url(item.youtube_id, item.query if item.query_type == "url" else None)
    user_name = user.username if user else None
    return ItemResponse(
        id=item.id,
        youtube_id=item.youtube_id,
        spotify_id=item.youtube_id or item.query,
        type="channel",
        query=item.query,
        query_type=item.query_type,
        name=item.name or item.query,
        youtube_url=url,
        spotify_url=url,
        image=item.image,
        banner_image=item.banner_image,
        video_count=item.video_count,
        subscriber_count=item.subscriber_count,
        view_count=item.view_count,
        view_count_delta=item.view_count_delta,
        delta_days=item.delta_days,
        followers=item.subscriber_count,
        playcount=item.view_count,
        track_count=item.video_count,
        followers_delta=None,
        playcount_delta=item.view_count_delta,
        track_count_delta=None,
        status=item.status,
        error_code=item.error_code,
        error_message=item.error_message,
        group=item.group,
        user_id=item.user_id,
        user_name=user_name,
        user_avatar=user.avatar if user else None,
        added_date=_format_added_date(item.created_at),
        last_checked=item.last_checked,
        created_at=item.created_at,
    )


def _with_item_filters(query, visible_ids: list[str], group: str | None = None, search: str | None = None):
    group = _plain_query_value(group)
    search = _plain_query_value(search)
    query = query.where(Item.user_id.in_(visible_ids))
    if group:
        query = query.where(Item.group == group)
    clean_search = (search or "").strip()
    if clean_search:
        like = f"%{clean_search.lower()}%"
        query = query.where(
            or_(
                func.lower(Item.name).like(like),
                func.lower(Item.youtube_id).like(like),
                func.lower(Item.query).like(like),
                func.lower(Item.group).like(like),
            )
        )
    return query


def _row_order_ids(user: User) -> list[str]:
    try:
        raw_preferences = getattr(user, "ui_preferences", None)
        preferences = json.loads(raw_preferences) if raw_preferences else {}
    except (json.JSONDecodeError, TypeError):
        return []
    row_order = preferences.get("row_order") if isinstance(preferences, dict) else []
    if not isinstance(row_order, list):
        return []
    return [str(value) for value in row_order if str(value).strip()][:5000]


def _sort_expression(sort: str | None):
    return {
        "name": func.lower(Item.name),
        "updated": func.coalesce(Item.last_checked, Item.created_at),
        "video": Item.video_count,
        "subscriber": Item.subscriber_count,
        "view": Item.view_count,
        "delta": Item.view_count_delta,
        "checked": Item.last_checked,
    }.get(sort or "")


def _apply_item_sort(query, current_user: User, sort: str | None, sort_direction: str | None):
    sort = _plain_query_value(sort)
    sort_direction = _plain_query_value(sort_direction, "asc")
    direction = "desc" if sort_direction == "desc" else "asc"
    expression = _sort_expression(sort)
    if expression is not None:
        ordered = expression.desc() if direction == "desc" else expression.asc()
        return query.order_by(ordered.nullslast(), Item.created_at.desc())

    row_order = _row_order_ids(current_user)
    if row_order:
        rank = case({item_id: idx for idx, item_id in enumerate(row_order)}, value=Item.id, else_=len(row_order))
        ordered = rank.desc() if direction == "desc" else rank.asc()
        return query.order_by(ordered, Item.created_at.desc())
    return query.order_by(Item.created_at.desc())


@router.get("/items", response_model=ItemListResponse)
async def list_items(
    group: str | None = Query(None),
    user_id: str | None = Query(None),
    search: str | None = Query(None),
    sort: str | None = Query(None),
    sort_direction: str | None = Query("asc"),
    limit: int = Query(5000, ge=1, le=10000),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    group = _plain_query_value(group)
    search = _plain_query_value(search)
    sort = _plain_query_value(sort)
    sort_direction = _plain_query_value(sort_direction, "asc")
    visible_ids = await _visible_user_ids(db, current_user, user_id)
    query = _with_item_filters(select(Item), visible_ids, group, search)
    count_query = _with_item_filters(select(func.count()).select_from(Item), visible_ids, group, search)
    query = _apply_item_sort(query, current_user, sort, sort_direction).offset(offset)
    if limit is not None:
        query = query.limit(limit)
    result = await db.execute(query)
    items = list(result.scalars().all())
    total = (await db.execute(count_query)).scalar() or 0
    users = await _user_map(db, list({item.user_id for item in items}))
    return ItemListResponse(items=[_response(item, users.get(item.user_id)) for item in items], total=total)


@router.get("/items/ids", response_model=ItemIdsResponse)
async def list_item_ids(
    group: str | None = Query(None),
    user_id: str | None = Query(None),
    search: str | None = Query(None),
    sort: str | None = Query(None),
    sort_direction: str | None = Query("asc"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    group = _plain_query_value(group)
    search = _plain_query_value(search)
    sort = _plain_query_value(sort)
    sort_direction = _plain_query_value(sort_direction, "asc")
    visible_ids = await _visible_user_ids(db, current_user, user_id)
    query = _with_item_filters(select(Item.id), visible_ids, group, search)
    count_query = _with_item_filters(select(func.count()).select_from(Item), visible_ids, group, search)
    result = await db.execute(_apply_item_sort(query, current_user, sort, sort_direction))
    total = (await db.execute(count_query)).scalar() or 0
    return ItemIdsResponse(ids=[str(row[0]) for row in result.all()], total=total)


@router.get("/items/summary", response_model=ItemSummaryResponse)
async def items_summary(
    group: str | None = Query(None),
    user_id: str | None = Query(None),
    search: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    group = _plain_query_value(group)
    search = _plain_query_value(search)
    visible_ids = await _visible_user_ids(db, current_user, user_id)
    all_total = (await db.execute(select(func.count()).select_from(Item).where(Item.user_id.in_(visible_ids)))).scalar() or 0

    filtered = _with_item_filters(select(Item.status), visible_ids, group, search).subquery()
    total = (await db.execute(select(func.count()).select_from(filtered))).scalar() or 0
    status_rows = (await db.execute(select(filtered.c.status, func.count()).group_by(filtered.c.status))).all()
    status_counts = {str(status or "").lower(): count for status, count in status_rows}

    group_rows = (
        await db.execute(
            select(Item.group, func.count())
            .where(Item.user_id.in_(visible_ids), Item.group.is_not(None), Item.group != "")
            .group_by(Item.group)
            .order_by(func.lower(Item.group))
        )
    ).all()
    return ItemSummaryResponse(
        total=total,
        all_total=all_total,
        active=status_counts.get("active", 0) + status_counts.get("completed", 0),
        errors=status_counts.get("error", 0) + status_counts.get("dead", 0),
        crawling=status_counts.get("crawling", 0) + status_counts.get("pending", 0),
        groups=[ItemGroupSummary(name=str(name), count=count) for name, count in group_rows if name],
    )


@router.post("/items/move")
async def move_items_group(
    payload: ItemMoveRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    visible_ids = await _visible_user_ids(db, current_user, payload.user_id)
    result = await db.execute(select(Item).where(Item.id.in_(payload.item_ids), Item.user_id.in_(visible_ids)))
    items = list(result.scalars().all())
    target_group = (payload.group or "").strip() or None
    for item in items:
        item.group = target_group
    return {"ok": True, "moved": len(items), "group": target_group}


@router.patch("/items/group")
async def rename_group(
    old_group: str = Query(...),
    new_group: str | None = Query(None),
    user_id: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    old_clean = old_group.strip()
    if not old_clean:
        raise HTTPException(status_code=400, detail="old_group is required")
    visible_ids = await _visible_user_ids(db, current_user, user_id)
    result = await db.execute(select(Item).where(Item.user_id.in_(visible_ids), Item.group == old_clean))
    items = list(result.scalars().all())
    next_group = (new_group or "").strip() or None
    for item in items:
        item.group = next_group
    return {"ok": True, "updated": len(items), "old_group": old_clean, "new_group": next_group}


@router.get("/items/{item_type}/{youtube_id}", response_model=ItemResponse)
async def get_item(
    item_type: str,
    youtube_id: str,
    user_id: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    visible_ids = await _visible_user_ids(db, current_user, user_id)
    result = await db.execute(
        select(Item).where(Item.user_id.in_(visible_ids), Item.youtube_id == youtube_id).order_by(Item.updated_at.desc())
    )
    item = result.scalars().first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    users = await _user_map(db, [item.user_id])
    return _response(item, users.get(item.user_id))


@router.delete("/items/{item_type}/{youtube_id}")
async def delete_item(
    item_type: str,
    youtube_id: str,
    user_id: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    visible_ids = await _visible_user_ids(db, current_user, user_id)
    result = await db.execute(select(Item).where(Item.user_id.in_(visible_ids), Item.youtube_id == youtube_id))
    item = result.scalars().first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    await db.execute(delete(CrawlJob).where(CrawlJob.item_id == item.id))
    await db.delete(item)
    return {"ok": True}


@router.delete("/items-by-id/{item_id}")
async def delete_item_by_id(
    item_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    visible_ids = await _visible_user_ids(db, current_user)
    result = await db.execute(select(Item).where(Item.id == item_id, Item.user_id.in_(visible_ids)))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    await db.execute(delete(CrawlJob).where(CrawlJob.item_id == item.id))
    await db.delete(item)
    return {"ok": True}


@router.delete("/items")
async def clear_items(
    group: str | None = Query(None),
    user_id: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    visible_ids = await _visible_user_ids(db, current_user, user_id)
    selected = select(Item.id).where(Item.user_id.in_(visible_ids))
    if group:
        selected = selected.where(Item.group == group)
    item_ids = [row[0] for row in (await db.execute(selected)).all()]
    if not item_ids:
        return {"ok": True, "deleted": 0, "group": group}
    await db.execute(delete(CrawlJob).where(CrawlJob.item_id.in_(item_ids)))
    result = await db.execute(delete(Item).where(Item.id.in_(item_ids)))
    return {"ok": True, "deleted": result.rowcount or 0, "group": group}


@router.post("/items/export")
async def export_items(payload: dict, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    item_ids = payload.get("item_ids") if isinstance(payload, dict) else []
    visible_ids = await _visible_user_ids(db, current_user)
    query = select(Item).where(Item.user_id.in_(visible_ids))
    if item_ids:
        query = query.where(Item.id.in_([str(x) for x in item_ids]))
    result = await db.execute(query.order_by(Item.created_at.desc()))
    users = await _user_map(db, visible_ids)
    rows = [_response(item, users.get(item.user_id)).model_dump(mode="json") for item in result.scalars().all()]
    if item_ids:
        order = {str(item_id): index for index, item_id in enumerate(item_ids)}
        rows.sort(key=lambda row: order.get(str(row.get("id")), len(order)))
    if payload.get("format") == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Channels"
        columns = [
            ("Tên", "name"),
            ("URL", "youtube_url"),
            ("Lần cuối", "last_checked"),
            ("Video", "video_count"),
            ("Subscriber", "subscriber_count"),
            ("View", "view_count"),
            ("Biến động / Ngày", "view_count_delta"),
            ("Group", "group"),
            ("Owner", "user_name"),
            ("Status", "status"),
        ]
        for column_index, (label, _) in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=column_index, value=label)
            cell.font = Font(bold=True)
        for row_index, row in enumerate(rows, start=2):
            for column_index, (_, key) in enumerate(columns, start=1):
                sheet.cell(row=row_index, column=column_index, value=row.get(key) or "")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_index, (label, key) in enumerate(columns, start=1):
            width = max(len(label), *(len(str(row.get(key) or "")) for row in rows[:200]))
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max(width + 2, 12), 42)
        output = io.BytesIO()
        workbook.save(output)
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="channels.xlsx"'},
        )
    return {"rows": rows, "items": rows, "count": len(rows)}
